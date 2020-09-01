# This program is mainly work with Excel, import the data in the Excel file into Python, and find the optimal path according to the needs of customers.
# To run this program propertly, you need to save the original excel file properly on your computer, for PC, the position should be on your desktop, for Mac, the position is the position of this excel file.
# You can start running this program with inputting a start city in the input box at the bottom of the program box (there is an inputting instruction )

#Input data
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
#  Open or modify your initial excel form following the inputting guidance.
#  Save your data chart in the specific saving path: for PC, the position should be on your desktop, for Mac, the position is the position of this excel file.

try:
    loc = 'c:/Users/User/Desktop/TexasAirways.xlsx' 
    wb = load_workbook(loc)  #Opens the file for Windows system
except:
    loc = 'TexasAirways.xlsx'
    wb = load_workbook(loc)  #Opens the file for Mac system
ws=wb['Sheet1']
import pandas as pd
import numpy as np

cityn=0
while ws.cell(row=1,column=cityn+2).value!= None:
      cityn=cityn+1
print('The amount of cities is ',cityn)  #Get the overall city numbers

distance=np.zeros((cityn,cityn),dtype=int)  #Intialize the distance with zero for any two cities
city=[]
for i in range(cityn):
    city+=[ws.cell(row=1,column=i+2).value]
D=pd.DataFrame(distance, index=city, columns=city) #Create the panda dataframe of distance 

for i in range(cityn):
    for j in range(cityn):
        if i==j:
            D.iloc[i,j]=0
        elif ws.cell(row=i+2,column=j+2).value==None and ws.cell(row=j+2,column=i+2).value==None:
            D.iloc[i,j]=float('inf')                         #In the case when there is no nonstop flight between two cities, return infinite to the distance dataframe
        elif ws.cell(row=i+2,column=j+2).value!=None and ws.cell(row=j+2,column=i+2).value==None:
            D.iloc[i,j]=ws.cell(row=i+2,column=j+2).value
        else:
            D.iloc[i,j]=ws.cell(row=j+2,column=i+2).value     #Input the data from excel to the distance dataframe created

List=range(0,cityn)

p=np.zeros((cityn,cityn))      #Create dataframe for path and initialize the path with 0
for i in range(0,cityn):
    for j in range(0,cityn):
        if i==j:
            p[i,j]=i       #For the path from city i to itself, return i in the dataframe
        if i!=j and D.iloc[i,j]<float('inf'):
            p[i,j]=j       #For the path from city i to city j, if there is a flight available, return j
        else:
            p[i,j]=-1     #For the path from city i to city j, if there is no flight available, return  -1
P=pd.DataFrame(p,index=List,columns=List)

Stop=' ' 
while Stop!='Stop': # The process is repeated if not entering "Stop"
    m=-1
    n=-1
    while m==-1 or n==-1 or m==n:
        m=-1
        n=-1
        for i in range(cityn-1): 
            print(str(city[i][0]),end=':')
            print(str(city[i]),end=', ')
        print(str(city[cityn-1][0]),end=':')
        print(str(city[cityn-1]),end='.')
        print('\n')
        ST=input('Enter the start place, e.g. Dallas or D: ')
        EN=input('Enter the destination, e.g. Houston or H: ')
        for i in range(cityn):
            if ST==ws.cell(row=1,column=i+2).value or ST==ws.cell(row=1,column=i+2).value[0]:
                n=i
                break  #Find the corresponding number of the start and end city

        for j in range(cityn):
            if EN==ws.cell(row=1,column=j+2).value or EN==ws.cell(row=1,column=j+2).value[0]:
                m=j
                break

        if m==-1 or n==-1:
            print('Inputting place is invalid! Input again!')
        elif m==n:
            print('Two place are same! Input again!')
        else:
            print('Good input!')

    for k in range(0,cityn):
        for i in range(0,cityn):
            for j in range(0,cityn):
                if k!=i and k!=j and D.iloc[i,j]>D.iloc[i,k]+D.iloc[k,j]:
                    D.iloc[i,j]=D.iloc[i,k]+D.iloc[k,j]  
                    #If the distance from city i to city i is larger than the distance from city i to city k plus the distance from city k to city j,
                    #replace the distance with the latter one in the distance dataframe
                    P.iloc[i,j]=P.iloc[i,k]  
                    #Return the path from city i to city j with the path from city i to city k as well
    
    print('\n','The minimum cost is',D.iloc[n,m])  #Show the minimum cost to the user
    
    if D.iloc[n,m]>99999:      #For the case when there are no flights which can connect city n and city m
        print('There is no route between',str(city[n]),' and ',str(city[m]))
    else: 
        path=[n]
        while n!=m:
            n=int(P.iloc[n,m])  #Find the least-cost path with the corresponding number of cities
            path=path+[n]
            Route=[]     

        for i in path:        #Convert the path with corresponding numbers to the path with city names
            Route=Route+[city[i]]

        print(' The optimal route is ',end=' ') #Show the path to the user
        for i in range(len(Route)-1): 
            print(str(Route[i]),end=' -- ')
        print(str(Route[len(Route)-1]),end='.')

        Stop=input('Enter "Stop" if you want to stop; otherwise, enter anything except "Stop".')
        #If the user don't want to move on to get the least-cost sequence of flights connecting other cities, 
        #he should enter "Stop" to quit; otherwise, continue.
        print('\n')
        
# The program requests you to input the start city and end city in the system. To improve the usability of the customers, you can input both the full name of the city or you can just input to capital letters of the city 
# To make the program more user-friendly, you can input the start city and the end city as many times as you want, the system will in the loop. If you want to quiet the system, you can input the “Stop” so that input boxes will stop to bump up. And the system will stop to run. Input anything other than “Stop”, the program will continue.


